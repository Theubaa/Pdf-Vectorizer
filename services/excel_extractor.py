from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import load_workbook


def extract_excel_blocks(
    raw_bytes: bytes,
    file_name: str,
) -> List[Dict[str, Any]]:
    """
    Convert an Excel workbook into semantic text blocks.

    Rules:
    - Process each sheet separately.
    - First non-empty row in a sheet is treated as the header.
    - Each subsequent row becomes one semantic block including:
        - sheet name
        - row number (1-based, as in Excel)
        - column names and values
    """
    wb = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)

    blocks: List[Dict[str, Any]] = []
    block_id = 0

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        # Determine header row (first non-empty row)
        header = None
        header_row_index = None
        for idx, row in enumerate(rows, start=1):
            if any(cell is not None and str(cell).strip() for cell in row):
                header = [str(c).strip() if c is not None else "" for c in row]
                header_row_index = idx
                break

        if header is None:
            continue

        # Process data rows after header
        for excel_row_index in range(header_row_index + 1, len(rows) + 1):
            row = rows[excel_row_index - 1]
            if row is None or not any(cell is not None and str(cell).strip() for cell in row):
                continue

            lines = [f"Sheet {sheet.title}, Row {excel_row_index} of {file_name}"]

            for col_name, value in zip(header, row):
                col_name = col_name or "column"
                value_str = "" if value is None else str(value).strip()
                lines.append(f"{col_name}: {value_str}")

            text_block = "\n".join(lines)
            blocks.append(
                {
                    "source_file": file_name,
                    "source_type": "excel",
                    "block_id": block_id,
                    "text": text_block,
                }
            )
            block_id += 1

    return blocks


